import openpyxl
from openpyxl.styles import Alignment,Font
from openpyxl.styles.borders import Border,Side
from openpyxl import Workbook

def cal_sheet(file_name):
    workbook=openpyxl.load_workbook(file_name)

    sheet = workbook['Midsem']
   
    for row in range(1,10):
        if(str(sheet[f'A{row}'].value)=="COs"):
            coRow = row
            break    

    for row in range(1,10):
        if (str(sheet[f'A{row}'].value)).startswith("Number of Students ="):
            total_roll=int((sheet[f'A{row}'].value).split('=')[-1].strip())
            
            break 

    target=((sheet[f'A{coRow+total_roll+3}'].value).split('>=')[-1].strip()).replace('%)', '')
    try:
        target = float(target)
        if target.is_integer():
            target = int(target)
    except ValueError:
        raise ValueError(f"Invalid target value: {target}")
    
    coRow =int(coRow) + 1  # Gives the start rowIndex for RollNo.
   
    for col in range (2,16):
        column_letter = openpyxl.utils.get_column_letter(col)
        sheet[f'{column_letter}{total_roll+coRow}']=f'=COUNT({column_letter}{coRow}:{column_letter}{total_roll+coRow-1})'

    for row in range (coRow,total_roll+coRow):    
        sheet[f'H{row}']=f'=ROUND(SUM(B{row}:G{row}),0)'

    for row in range (coRow,total_roll+coRow):    
        sheet[f'K{row}']=f'=ROUND(SUM(I{row}:J{row}),0)'
        
    for row in range (coRow,total_roll+coRow):    
        sheet[f'N{row}']=f'=ROUND(SUM(L{row}:M{row}),0)'
            
    for row in range (coRow,total_roll+coRow):    
        sheet[f'O{row}']=f'=ROUND(SUM(H{row},K{row},N{row}),0)'

    for col in range (2,16):
        column_letter = openpyxl.utils.get_column_letter(col)
        sheet[f'{column_letter}{total_roll+coRow+1}']=f'=ROUND(AVERAGE({column_letter}{coRow}:{column_letter}{total_roll+coRow-1}),0)'
    
    for col in range (2,8):
        column_letter = openpyxl.utils.get_column_letter(col)
        sheet[f'{column_letter}{total_roll+coRow+2}']=f'=COUNTIF({column_letter}{coRow}:{column_letter}{total_roll+coRow-1},">={target/100 * 2}")'

    for col in range (8,16):
        column_letter = openpyxl.utils.get_column_letter(col)
        sheet[f'{column_letter}{total_roll+coRow+2}']=f'=COUNTIF({column_letter}{coRow}:{column_letter}{total_roll+coRow-1},">={target/100 * 5}")'

    for col in range (2,16):
        column_letter = openpyxl.utils.get_column_letter(col)
        sheet[f'{column_letter}{total_roll+coRow+3}']=f'=ROUND(({column_letter}{total_roll+coRow+2}/{column_letter}{total_roll+coRow})*100,1)'


    for col in range (2,16):
        column_letter = openpyxl.utils.get_column_letter(col)
        sheet[f'{column_letter}{total_roll+coRow+4}']=f'=COUNTIF({column_letter}{coRow}:{column_letter}{total_roll+coRow-1},">="&{column_letter}{total_roll+coRow+1})'

    for col in range (2,16):
        column_letter = openpyxl.utils.get_column_letter(col)
        sheet[f'{column_letter}{total_roll+coRow+5}']=f'=ROUND(({column_letter}{total_roll+coRow+4}/{column_letter}{total_roll+coRow})*100,1)'

    for col in range (2,16):
        column_letter = openpyxl.utils.get_column_letter(col)
        sheet[f'{column_letter}{total_roll+coRow+6}']=f'=IF({column_letter}{total_roll+coRow+3}<60,1,IF(AND({column_letter}{total_roll+coRow+3}>59,{column_letter}{total_roll+coRow+3}<70),2,IF(AND({column_letter}{total_roll+coRow+3}>69,{column_letter}{total_roll+coRow+3}<80),3,4)))'
        

    coTableRow = total_roll+coRow+9

    columns_with_1 = []
    columns_with_2 = []
    columns_with_3 = []
    columns_with_4 = []
    columns_with_5 = []
    columns_with_6 = [] 

        
    column_range = ['B', 'C', 'D', 'E', 'F', 'G', 'I', 'J', 'L', 'M']

    # Iterate through cells in the specified column range
    for column_letter in column_range:
        # Get the cell in row 2 corresponding to the column letter
        cell = sheet[f"{column_letter}{coRow-1}"]
        # Check if the cell has a value
        if cell.value :
            # values = [int(val.strip()) for val in str(cell.value).split(',') if val.strip().isdigit()]  #This for without CO like 1,2,3
            values = [int(val.strip()) for val in str(cell.value)[2:].split(',') if val.strip().isdigit()]   #This for with CO like CO1,2,3
            # Check if '1' is present in the list of values then 2 and 3,4and 5
            # print(values)
            if 1 in values:
                # Add the cell value to the list for calculation
                columns_with_1.append(column_letter)
            if 2 in values:
                # Add the cell value to the list for calculation
                columns_with_2.append(column_letter)
            if 3 in values:
                # Add the cell value to the list for calculation
                columns_with_3.append(column_letter)
            if 4 in values:
                # Add the cell value to the list for calculation
                columns_with_4.append(column_letter)
            if 5 in values:
                # Add the cell value to the list for calculation
                columns_with_5.append(column_letter)
            if 6 in values:
                columns_with_6.append(column_letter)

    # Calculate the average using Excel formula
    if columns_with_1:
        average_formula = f"=ROUND(AVERAGE({','.join([f'{col_letter}{total_roll+coRow+6}' for col_letter in columns_with_1])}),1)"
        sheet[f'G{coTableRow+1}'] = average_formula
    else:
        sheet[f'G{coTableRow+1}'] = '-'

        
    if columns_with_2:
        average_formula = f"=ROUND(AVERAGE({','.join([f'{col_letter}{total_roll+coRow+6}' for col_letter in columns_with_2])}),1)"
        sheet[f'G{coTableRow+2}'] = average_formula
    else:
        sheet[f'G{coTableRow+2}'] = '-'
    
    if columns_with_3:
        average_formula = f"=ROUND(AVERAGE({','.join([f'{col_letter}{total_roll+coRow+6}' for col_letter in columns_with_3])}),1)"
        sheet[f'G{coTableRow+3}'] = average_formula
    else:
        sheet[f'G{coTableRow+3}'] = '-'
        
    if columns_with_4:
        average_formula = f"=ROUND(AVERAGE({','.join([f'{col_letter}{total_roll+coRow+6}' for col_letter in columns_with_4])}),1)"
        sheet[f'G{coTableRow+4}'] = average_formula
    else:
        sheet[f'G{coTableRow+4}'] = '-'

    if columns_with_5:
        average_formula = f"=ROUND(AVERAGE({','.join([f'{col_letter}{total_roll+coRow+6}' for col_letter in columns_with_5])}),1)"
        sheet[f'G{coTableRow+5}'] = average_formula
    else:
        sheet[f'G{coTableRow+5}'] = '-'

    if columns_with_6:
        average_formula = f"=ROUND(AVERAGE({','.join([f'{col_letter}{total_roll+coRow+6}' for col_letter in columns_with_6])}),1)"
        sheet[f'G{coTableRow+6}'] = average_formula
    else:
        sheet[f'G{coTableRow+6}'] = '-'
    
    map_midsem_co_arr=[f"=Midsem!G{coTableRow+1}",f"=Midsem!G{coTableRow+2}",f"=Midsem!G{coTableRow+3}",f"=Midsem!G{coTableRow+4}",f"=Midsem!G{coTableRow+5}",f"=Midsem!G{coTableRow+6}"]
   #<----------------------End Sem------------------->
   
    sheet1 = workbook['Endsem']

    endCol='B'
    
    sheet1[f'B{total_roll+4}'] = f'=COUNT({endCol}4:{endCol}{total_roll+3})'
    sheet1[f'B{total_roll+5}'] = f'=ROUND(AVERAGE({endCol}4:{endCol}{total_roll+3}), 0)'
    sheet1[f'B{total_roll+6}'] = f'=COUNTIF({endCol}4:{endCol}{total_roll+3}, ">={target / 100 * 80}")'
    sheet1[f'B{total_roll+7}'] = f'=ROUND({sheet1[f"B{str(total_roll+6)}"].coordinate} / {sheet1[f"B{str(total_roll+4)}"].coordinate} * 100, 1)'
    sheet1[f'B{total_roll+8}'] = f'=COUNTIF({endCol}4:{endCol}{total_roll+3}, ">="&{endCol}{total_roll+5})'
    sheet1[f'B{total_roll+9}'] = f'=ROUND({sheet1[f"B{str(total_roll+8)}"].coordinate} / {sheet1[f"B{str(total_roll+4)}"].coordinate} * 100, 1)'
    sheet1[f'B{total_roll+10}'] = f'=IF({sheet1[f"B{str(total_roll+7)}"].coordinate}<60, 1, IF(AND({sheet1[f"B{str(total_roll+7)}"].coordinate}>59, {sheet1[f"B{str(total_roll+7)}"].coordinate}<70), 2, IF(AND({sheet1[f"B{str(total_roll+7)}"].coordinate}>69, {sheet1[f"B{str(total_roll+7)}"].coordinate}<80), 3, 4)))'
    
        
    check = [int(val.strip()) for val in str(sheet1['B3'].value)[2:].split(',') if val.strip().isdigit()]
    
    if 1 in check:
        sheet1[f'B{total_roll+14}']=sheet1[f'B{total_roll+10}'].value
    else :
        sheet1[f'B{total_roll+14}']="-"
    
    if 2 in check:
        sheet1[f'B{total_roll+15}']=sheet1[f'B{total_roll+10}'].value
    else :
        sheet1[f'B{total_roll+15}']="-"
        
    if 3 in check:
        sheet1[f'B{total_roll+16}']=sheet1[f'B{total_roll+10}'].value
    else :
        sheet1[f'B{total_roll+16}']="-"
        
    if 4 in check:
        sheet1[f'B{total_roll+17}']=sheet1[f'B{total_roll+10}'].value
    else :
        sheet1[f'B{total_roll+17}']="-"
        
    if 5 in check:
        sheet1[f'B{total_roll+18}']=sheet1[f'B{total_roll+10}'].value
    else :
        sheet1[f'B{total_roll+18}']="-"
        
    if 6 in check:
        sheet1[f'B{total_roll+19}']=sheet1[f'B{total_roll+10}'].value
    else :
        sheet1[f'B{total_roll+19}']="-"
    
    map_endsem_co_arr=[f'=Endsem!B{total_roll+14}',f'=Endsem!B{total_roll+15}',f'=Endsem!B{total_roll+16}',f'=Endsem!B{total_roll+17}',f'=Endsem!B{total_roll+18}',f'=Endsem!B{total_roll+19}']
    #<----------------------CA------------------->
    
    sheet2=workbook['CA']
    
    
    for col in ['B','C','D']:
        
        sheet2[f'{col}{total_roll+6}'] = f'=COUNT({col}6:{col}{total_roll+5})'
        sheet2[f'{col}{total_roll+7}'] = f'=ROUND(AVERAGE({col}6:{col}{total_roll+5}), 0)'
        sheet2[f'{col}{total_roll+8}'] = f'=COUNTIF({col}6:{col}{total_roll+5}, ">={target / 100 * 10}")'
        sheet2[f'{col}{total_roll+9}'] = f'=ROUND({sheet2[f"{col}{total_roll+8}"].coordinate} / {sheet2[f"{col}{total_roll+6}"].coordinate} * 100, 1)'
        sheet2[f'{col}{total_roll+10}'] = f'=COUNTIF({col}6:{col}{total_roll+5}, ">="&{col}{total_roll+7})'
        sheet2[f'{col}{total_roll+11}'] = f'=ROUND({sheet2[f"{col}{total_roll+10}"].coordinate} / {sheet2[f"{col}{total_roll+6}"].coordinate} * 100, 1)'
        sheet2[f'{col}{total_roll+12}'] = f'=IF({sheet2[f"{col}{total_roll+9}"].coordinate}<60, 1, IF(AND({sheet2[f"{col}{total_roll+9}"].coordinate}>59, {sheet2[f"{col}{total_roll+9}"].coordinate}<70), 2, IF(AND({sheet2[f"{col}{total_roll+9}"].coordinate}>69, {sheet2[f"{col}{total_roll+9}"].coordinate}<80), 3, 4)))'
    
    
    columns_CA_1 = []
    columns_CA_2 = []
    columns_CA_3 = []
    columns_CA_4 = []
    columns_CA_5 = []
    columns_CA_6 = [] 
    
    for col in ['B','C','D']:
        cell = sheet2[f"{col}5"]
        if cell.value :
            value2 = [int(val.strip()) for val in str(cell.value)[2:].split(',') if val.strip().isdigit()]   #This for with CO like CO1,2,3
            if 1 in value2:
                # Add the cell value to the list for calculation
                columns_CA_1.append(col)
            if 2 in value2:
                # Add the cell value to the list for calculation
                columns_CA_2.append(col)
            if 3 in value2:
                # Add the cell value to the list for calculation
                columns_CA_3.append(col)
            if 4 in value2:
                # Add the cell value to the list for calculation
                columns_CA_4.append(col)
            if 5 in value2:
                # Add the cell value to the list for calculation
                columns_CA_5.append(col)
            if 6 in value2:
                columns_CA_6.append(col)

    # Calculate the average using Excel formula
    if columns_CA_1:
        average_formula = f"=ROUND(AVERAGE({','.join([f'{col_letter}{total_roll+12}' for col_letter in columns_CA_1])}),1)"
        sheet2[f'C{total_roll+16}'] = average_formula
    else:
        sheet2[f'C{total_roll+16}'] = '-'

        
    if columns_CA_2:
        average_formula = f"=ROUND(AVERAGE({','.join([f'{col_letter}{total_roll+12}' for col_letter in columns_CA_2])}),1)"
        sheet2[f'C{total_roll+17}'] = average_formula
    else:
        sheet2[f'C{total_roll+17}'] = '-'
    
    if columns_CA_3:
        average_formula = f"=ROUND(AVERAGE({','.join([f'{col_letter}{total_roll+12}' for col_letter in columns_CA_3])}),1)"
        sheet2[f'C{total_roll+18}'] = average_formula
    else:
        sheet2[f'C{total_roll+18}'] = '-'
        
    if columns_CA_4:
        average_formula = f"=ROUND(AVERAGE({','.join([f'{col_letter}{total_roll+12}' for col_letter in columns_CA_4])}),1)"
        sheet2[f'C{total_roll+19}'] = average_formula
    else:
        sheet2[f'C{total_roll+19}'] = '-'

    if columns_CA_5:
        average_formula = f"=ROUND(AVERAGE({','.join([f'{col_letter}{total_roll+12}' for col_letter in columns_CA_5])}),1)"
        sheet2[f'C{total_roll+20}'] = average_formula
    else:
        sheet2[f'C{total_roll+20}'] = '-'

    if columns_CA_6:
        average_formula = f"=ROUND(AVERAGE({','.join([f'{col_letter}{total_roll+12}' for col_letter in columns_CA_6])}),1)"
        sheet2[f'C{total_roll+21}'] = average_formula
    else:
        sheet2[f'C{total_roll+21}'] = '-'
    
    map_CA_co_arr=[f'=CA!C{total_roll+16}',f'=CA!C{total_roll+17}',f'=CA!C{total_roll+18}',f'=CA!C{total_roll+19}',f'=CA!C{total_roll+20}',f'=CA!C{total_roll+21}']
    
    #<----------------------Quiz------------------->
    sheet3=workbook['Quiz']
    
    for col in ['C', 'D', 'E', 'F', 'G','H','I','J','K','L']:
        
        sheet3[f'{col}{total_roll+3}'] = f'=COUNT({col}3:{col}{total_roll+2})'
        sheet3[f'{col}{total_roll+4}'] = f'=ROUND(AVERAGE({col}3:{col}{total_roll+2}), 0)'
        sheet3[f'{col}{total_roll+5}'] = f'=COUNTIF({col}3:{col}{total_roll+2}, ">={target / 100 * 2}")'
        sheet3[f'{col}{total_roll+6}'] = f'=ROUND({sheet3[f"{col}{total_roll+5}"].coordinate} / {sheet3[f"{col}{total_roll+3}"].coordinate} * 100, 1)'
        sheet3[f'{col}{total_roll+7}'] = f'=COUNTIF({col}3:{col}{total_roll+2}, ">="&{col}{total_roll+4})'
        sheet3[f'{col}{total_roll+8}'] = f'=ROUND({sheet3[f"{col}{total_roll+7}"].coordinate} / {sheet3[f"{col}{total_roll+3}"].coordinate} * 100, 1)'
        sheet3[f'{col}{total_roll+9}'] = f'=IF({sheet3[f"{col}{total_roll+6}"].coordinate}<60, 1, IF(AND({sheet3[f"{col}{total_roll+6}"].coordinate}>59, {sheet3[f"{col}{total_roll+6}"].coordinate}<70), 2, IF(AND({sheet3[f"{col}{total_roll+6}"].coordinate}>69, {sheet3[f"{col}{total_roll+6}"].coordinate}<80), 3, 4)))'
    
    columns_QZ_1 = []
    columns_QZ_2 = []
    columns_QZ_3 = []
    columns_QZ_4 = []
    columns_QZ_5 = []
    columns_QZ_6 = []
    
    for col in ['C', 'D', 'E', 'F', 'G','H', 'I', 'J','K','L']:
        cell = sheet3[f"{col}2"]
        if cell.value :
            value3 = [int(val.strip()) for val in str(cell.value)[2:].split(',') if val.strip().isdigit()]   #This for with CO like CO1,2,3
            if 1 in value3:
                # Add the cell value to the list for calculation
                columns_QZ_1.append(col)
            if 2 in value3:
                # Add the cell value to the list for calculation
                columns_QZ_2.append(col)
            if 3 in value3:
                # Add the cell value to the list for calculation
                columns_QZ_3.append(col)
            if 4 in value3:
                # Add the cell value to the list for calculation
                columns_QZ_4.append(col)
            if 5 in value3:
                # Add the cell value to the list for calculation
                columns_QZ_5.append(col)
            if 6 in value3:
                columns_QZ_6.append(col)

    # Calculate the average using Excel formula
    if columns_QZ_1:
        average_formula = f"=ROUND(AVERAGE({','.join([f'{col_letter}{total_roll+9}' for col_letter in columns_QZ_1])}),1)"
        sheet3[f'D{total_roll+13}'] = average_formula
    else:
        sheet3[f'D{total_roll+13}'] = '-'

        
    if columns_QZ_2:
        average_formula = f"=ROUND(AVERAGE({','.join([f'{col_letter}{total_roll+9}' for col_letter in columns_QZ_2])}),1)"
        sheet3[f'D{total_roll+14}'] = average_formula
    else:
        sheet3[f'D{total_roll+14}'] = '-'
    
    if columns_QZ_3:
        average_formula = f"=ROUND(AVERAGE({','.join([f'{col_letter}{total_roll+9}' for col_letter in columns_QZ_3])}),1)"
        sheet3[f'D{total_roll+15}'] = average_formula
    else:
        sheet3[f'D{total_roll+15}'] = '-'
        
    if columns_QZ_4:
        average_formula = f"=ROUND(AVERAGE({','.join([f'{col_letter}{total_roll+9}' for col_letter in columns_QZ_4])}),1)"
        sheet3[f'D{total_roll+16}'] = average_formula
    else:
        sheet3[f'D{total_roll+16}'] = '-'

    if columns_QZ_5:
        average_formula = f"=ROUND(AVERAGE({','.join([f'{col_letter}{total_roll+9}' for col_letter in columns_QZ_5])}),1)"
        sheet3[f'D{total_roll+17}'] = average_formula
    else:
        sheet3[f'D{total_roll+17}'] = '-'

    if columns_QZ_6:
        average_formula = f"=ROUND(AVERAGE({','.join([f'{col_letter}{total_roll+9}' for col_letter in columns_QZ_6])}),1)"
        sheet3[f'D{total_roll+18}'] = average_formula
    else:
        sheet3[f'D{total_roll+18}'] = '-'
    
    
    map_quiz_co_arr=[f'=Quiz!D{total_roll+13}',f'=Quiz!D{total_roll+14}',f'=Quiz!D{total_roll+15}',f'=Quiz!D{total_roll+16}',f'=Quiz!D{total_roll+17}',f'=Quiz!D{total_roll+18}']
    
    
    #<----------------------Survey------------------->
    
    sheet4=workbook['Survey']
    
    for col in ['G','H','I','J','K','L']:
        
        sheet4[f'{col}{total_roll+4}'] = f'=COUNT({col}2:{col}{total_roll+1})'
        sheet4[f'{col}{total_roll+5}'] = f'=COUNTIF({col}2:{col}{total_roll+1}, ">=4")'
        sheet4[f'{col}{total_roll+6}'] = f'=ROUND(({col}{total_roll+5}/{col}{total_roll+4}*100), 1)'
        sheet4[f'{col}{total_roll+8}'] = f'=IF({col}{total_roll+6}<60,1,IF(AND({col}{total_roll+6}>59,{col}{total_roll+6}<70),2,IF(AND({col}{total_roll+6}>69,{col}{total_roll+6}<80),3,4)))'
        
    map_survey_co_arr=[f'=Survey!G{total_roll+8}',f'=Survey!H{total_roll+8}',f'=Survey!I{total_roll+8}',f'=Survey!J{total_roll+8}',f'=Survey!K{total_roll+8}',f'=Survey!L{total_roll+8}']
    
    #<----------------------Attainment------------------->
    sheet5=workbook['Attainment']
    
    for i in range(0,6):
        sheet5[f'B{33+i}']=map_midsem_co_arr[i]
    
    for i in range(0,6):
        sheet5[f'C{33+i}']=map_CA_co_arr[i]
        
    for i in range(0,6):
        sheet5[f'D{33+i}']=map_quiz_co_arr[i]
        
    for i in range(0,6):
        sheet5[f'E{33+i}']=map_endsem_co_arr[i]
        
    for i in range(0,6):
        sheet5[f'G{33+i}']=map_survey_co_arr[i]
        
    for i in range(0,6):
        sheet5[f'E{43+i}']=map_survey_co_arr[i] 
        
    for i in range(0,6):
        sheet5[f'F{33+i}']=f'=ROUND(0.7*E{33+i}+0.3*(AVERAGE(B{33+i},C{33+i},D{33+i})),1)'
           
    for i in range(0,6):
        sheet5[f'D{43+i}']=sheet5[f'F{33+i}'].value        
         
         
    workbook.save('C:/Users/saira/Downloads/Calculated.xlsx')
    
    
    
    
    
#cal_sheet()