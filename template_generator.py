import openpyxl
from openpyxl.styles import Alignment,Font
from openpyxl.styles.borders import Border,Side
from openpyxl import Workbook

def template_gen(basic_values_temp,midSem_Co_values_temp,CA1_Co_arr_temp,CA2_Co_arr_temp,CA3_Co_arr_temp):
    workbook=Workbook() 
    print("Number of CAs:", basic_values_temp[10])
    sheet1 = workbook.active
    sheet1.title = "Midsem"

    sheet2=workbook.create_sheet(title="Endsem")
    sheet3=workbook.create_sheet(title="CA1")
    sheet4=workbook.create_sheet(title="CA2")
    if basic_values_temp[10]=="3":
        sheet7=workbook.create_sheet(title="CA3")
        print("Created")
    sheet5=workbook.create_sheet(title="Survey")
    sheet6=workbook.create_sheet(title='Attainment')
    
    
    sheet1.column_dimensions['A'].width = 42
    sheet1.merge_cells("A1:O1")
    sheet1["A1"].value="Vivekanand Education Society's Institute of Technology"

    
    sheet1.merge_cells("A2:O2")
    sheet1["A2"].value="Department of "+basic_values_temp[1]+""

    sheet1.merge_cells("A3:O3")
    sheet1["A3"].value="Academic Year :"+basic_values_temp[5]+""
    
    sheet1.merge_cells("A4:O4")
    
    sheet1.merge_cells("A5:O5")
    sheet1["A5"].value="  Subject : "+basic_values_temp[4]+"                                                                                                                                                                       Class : "+basic_values_temp[7]+""


    sheet1.merge_cells("A6:O6")
    sheet1["A6"].value="  Subject Teacher :"+basic_values_temp[6]+"                                                                                                                                                                Semester : "+basic_values_temp[3]+""
  
    

    sheet1.merge_cells("A7:O7")
    sheet1["A7"].value="Number of Students ="+basic_values_temp[0]+""

    sheet1["A8"]="Roll Nos."
    sheet1["B8"]="1a"
    sheet1["C8"]="1b"
    sheet1["D8"]="1c"
    sheet1["E8"]="1d"
    sheet1["F8"]="1e"
    sheet1["G8"]="1f"
    sheet1["H8"]="Q1"
    sheet1["I8"]="2a"
    sheet1["J8"]="2b"
    sheet1["K8"]="Q2"
    sheet1["L8"]="3a"
    sheet1["M8"]="3b"
    sheet1["N8"]="Q3"
    sheet1["O8"]="Total"

    sheet1["A9"]="COs"
    sheet1["B9"]="CO"+midSem_Co_values_temp[0]+""
    sheet1["C9"]="CO"+midSem_Co_values_temp[1]+""
    sheet1["D9"]="CO"+midSem_Co_values_temp[2]+""
    sheet1["E9"]="CO"+midSem_Co_values_temp[3]+""
    sheet1["F9"]="CO"+midSem_Co_values_temp[4]+""
    sheet1["G9"]="CO"+midSem_Co_values_temp[5]+""
    # sheet1["H9"]=""
    sheet1["I9"]="CO"+midSem_Co_values_temp[6]+""
    sheet1["J9"]="CO"+midSem_Co_values_temp[7]+""
    # sheet1["K9"]=""
    sheet1["L9"]="CO"+midSem_Co_values_temp[8]+""
    sheet1["M9"]="CO"+midSem_Co_values_temp[9]+""
    # sheet1["N9"]=""
    sheet1["O9"]=20

    total_roll=int(basic_values_temp[0])
    for i in range(1,total_roll+1):
        sheet1[f'A{i+9}']=i
        sheet1[f'A{i+9}'].alignment= Alignment(horizontal='center', vertical='center')
        for col in ['B', 'C', 'D', 'E', 'F', 'G','H','I', 'J','K','L','M','N','O']:
            sheet1[f'{col}{i+9}'].alignment= Alignment(horizontal='center', vertical='center')
        
    sheet1['A1'].alignment= Alignment(horizontal='center', vertical='center')
    sheet1['A2'].alignment= Alignment(horizontal='center', vertical='center')
    sheet1['A3'].alignment= Alignment(horizontal='center', vertical='center')
    sheet1['A4'].alignment= Alignment(horizontal='left', vertical='center')
    sheet1['A5'].alignment= Alignment(horizontal='left', vertical='center')
    sheet1['A6'].alignment= Alignment(horizontal='left', vertical='center')
    sheet1['A7'].alignment= Alignment(horizontal='left', vertical='center')

    
    
    column_range = ['A','B', 'C', 'D', 'E', 'F', 'G','H','I', 'J','K','L','M','N','O']
    for i in range(7,10):
        for col in column_range :
            sheet1[f'{col}{i}'].alignment= Alignment(horizontal='center', vertical='center')
    
    for i in range(total_roll+10,total_roll+17):
        for col in ['B', 'C', 'D', 'E', 'F', 'G','H','I', 'J','K','L','M','N','O'] :
            sheet1[f'{col}{i}'].alignment= Alignment(horizontal='center', vertical='center')

    for i in range(1,total_roll+17):
        for col in column_range:
            sheet1[f'{col}{i}'].border=Border(top=Side(style='thin',color='000000'),right=Side(style='thin',color='000000'),left=Side(style='thin',color='000000'),bottom=Side(style='thin',color='000000'))
    
    for col in column_range:
        sheet1[f'{col}{total_roll+16}'].font=Font(bold=True)
          
    for i in range(1,10):
        for col in column_range:
          sheet1[f'{col}{i}'].font=Font(bold=True)
    
    sheet1[f'A{total_roll+10}'] = 'Count(Attempted)'
    sheet1[f'A{total_roll+11}'] = 'Average Marks'
    sheet1[f'A{total_roll+12}'] = f'Count(>={basic_values_temp[9]}%)'
    sheet1[f'A{total_roll+13}'] = f'% Count(>={basic_values_temp[9]}% w.r.t appeared)'
    sheet1[f'A{total_roll+14}'] = 'Count(>=Average Marks of class)'
    sheet1[f'A{total_roll+15}']= "% Count(>=Average Marks of class w.r.t appeared)"
    sheet1[f'A{total_roll+16}'] = f'AL(Based on >={basic_values_temp[9]}% Count) (All COs)'
    
    sheet1[f'F{total_roll+19}'] = "COs"
    sheet1[f'F{total_roll+19}'].font=Font(bold=True)
    sheet1[f'G{total_roll+19}'] = "AL"
    sheet1[f'G{total_roll+19}'].font=Font(bold=True)
    sheet1[f'F{total_roll+20}'] = 'CO1'
    sheet1[f'F{total_roll+21}'] = 'CO2'
    sheet1[f'F{total_roll+22}'] = 'CO3'
    sheet1[f'F{total_roll+23}'] = 'CO4'
    sheet1[f'F{total_roll+24}'] = 'CO5'
    sheet1[f'F{total_roll+25}'] = 'CO6'

    for i in range(total_roll+19,total_roll+26):
        sheet1[f'F{i}'].border=Border(top=Side(style='thin',color='000000'),right=Side(style='thin',color='000000'),left=Side(style='thin',color='000000'),bottom=Side(style='thin',color='000000'))
        sheet1[f'G{i}'].border=Border(top=Side(style='thin',color='000000'),right=Side(style='thin',color='000000'),left=Side(style='thin',color='000000'),bottom=Side(style='thin',color='000000'))
        sheet1[f'F{i}'].alignment= Alignment(horizontal='center', vertical='center')
        sheet1[f'G{i}'].alignment= Alignment(horizontal='center', vertical='center')
        
    #<-----------End Semester Template------------->
    
    sheet2.column_dimensions['A'].width =42
    sheet2.column_dimensions['B'].width =22
    sheet2['A1']="Roll No."  
    sheet2['A1'].font=Font(bold=True) 
    sheet2['B1']="ESE(TH)"
    sheet2['B1'].font=Font(bold=True)     
    sheet2['B2']="ALL Qs"
    sheet2['B2'].font=Font(bold=True) 
    sheet2['B3']="CO"+basic_values_temp[8]+""
    sheet2['B3'].font=Font(bold=True) 
    
    for i in range(1,total_roll+1):
        sheet2[f'A{i+3}'] =i
       
    
    sheet2[f'A{total_roll+4}']="Count(Attempted)"
    sheet2[f'A{total_roll+5}']="Average Marks"
    
   
    sheet2[f'A{total_roll+6}']=f"Count(>={basic_values_temp[9]}%)"
    
    sheet2[f'A{total_roll+7}']=f"% Count(>={basic_values_temp[9]}% w.r.t appeared)"
    
    sheet2[f'A{total_roll+8}']="Count(>=Average Marks of class)"
    sheet2[f'A{total_roll+9}']="% Count(>=Average Marks of class w.r.t appeared)"
    
    sheet2[f'A{total_roll+10}']=f"AL(Based on >={basic_values_temp[9]}% Count) (All COs)"
    sheet2[f'A{total_roll+10}'].font=Font(bold=True) 
    sheet2[f'B{total_roll+10}'].font=Font(bold=True)
    
    for i in range(1,total_roll+11):
        sheet2[f'A{i}'].alignment= Alignment(horizontal='center', vertical='center')     
        sheet2[f'A{i}'].border=Border(top=Side(style='thin',color='000000'),right=Side(style='thin',color='000000'),left=Side(style='thin',color='000000'),bottom=Side(style='thin',color='000000'))
        sheet2[f'B{i}'].alignment= Alignment(horizontal='center', vertical='center')     
        sheet2[f'B{i}'].border=Border(top=Side(style='thin',color='000000'),right=Side(style='thin',color='000000'),left=Side(style='thin',color='000000'),bottom=Side(style='thin',color='000000'))
        if i>=total_roll+4 :
            sheet2[f'A{i}'].alignment= Alignment(horizontal='left', vertical='center') 
     
    sheet2[f'A{total_roll+13}'] = "COs"
    sheet2[f'A{total_roll+13}'].font=Font(bold=True)
    sheet2[f'B{total_roll+13}'] = "AL"
    sheet2[f'B{total_roll+13}'].font=Font(bold=True)
    sheet2[f'A{total_roll+14}'] = 'CO1'
    sheet2[f'A{total_roll+15}'] = 'CO2'
    sheet2[f'A{total_roll+16}'] = 'CO3'
    sheet2[f'A{total_roll+17}'] = 'CO4'
    sheet2[f'A{total_roll+18}'] = 'CO5'
    sheet2[f'A{total_roll+19}'] = 'CO6' 
    
    for i in range(total_roll+13,total_roll+20):
        sheet2[f'A{i}'].border=Border(top=Side(style='thin',color='000000'),right=Side(style='thin',color='000000'),left=Side(style='thin',color='000000'),bottom=Side(style='thin',color='000000'))
        sheet2[f'B{i}'].border=Border(top=Side(style='thin',color='000000'),right=Side(style='thin',color='000000'),left=Side(style='thin',color='000000'),bottom=Side(style='thin',color='000000'))
        sheet2[f'A{i}'].alignment= Alignment(horizontal='center', vertical='center')
        sheet2[f'B{i}'].alignment= Alignment(horizontal='center', vertical='center')
        
   
    
    def make_CA_Type_PPT_NPTEL(mysheet2,arr):
        mysheet2.column_dimensions['A'].width = 42
        mysheet2.column_dimensions['B'].width = 16
        mysheet2.merge_cells("A1:B1")
        mysheet2.merge_cells("A2:B2")
        mysheet2.merge_cells("A3:B3")
        mysheet2['A2']="CA Marksheet"
        mysheet2['A2'].font=Font(bold=True)
        mysheet2['A4']="Roll No."
        mysheet2['A4'].font=Font(bold=True)
        mysheet2['B4']="CA1"
        mysheet2['B4'].font=Font(bold=True)
        mysheet2['B5']="CO"+arr[0]+""
        mysheet2['B5'].font=Font(bold=True)
        
        for i in range(1 ,total_roll+1):
            mysheet2[f'A{i+5}']=i
        
        mysheet2[f'A{total_roll+6}']="Count(Attempted)"
        mysheet2[f'A{total_roll+7}']="Average Marks"
        
    
        mysheet2[f'A{total_roll+8}']=f"Count(>={basic_values_temp[9]}%)"
        
        
        mysheet2[f'A{total_roll+9}']=f"% Count(>={basic_values_temp[9]}% w.r.t appeared)"
        
        mysheet2[f'A{total_roll+10}']="Count(>=Average Marks of class)"
        mysheet2[f'A{total_roll+11}']="% Count(>=Average Marks of class w.r.t appeared)"
        
        mysheet2[f'A{total_roll+12}']=f"AL(Based on >={basic_values_temp[9]}% Count) (All COs)"
        mysheet2[f'A{total_roll+12}'].font=Font(bold=True)
        
        for i in range(1,total_roll+13):
            mysheet2[f'A{i}'].alignment= Alignment(horizontal='center', vertical='center')     
            mysheet2[f'A{i}'].border=Border(top=Side(style='thin',color='000000'),right=Side(style='thin',color='000000'),left=Side(style='thin',color='000000'),bottom=Side(style='thin',color='000000'))
            mysheet2[f'B{i}'].alignment= Alignment(horizontal='center', vertical='center')     
            mysheet2[f'B{i}'].border=Border(top=Side(style='thin',color='000000'),right=Side(style='thin',color='000000'),left=Side(style='thin',color='000000'),bottom=Side(style='thin',color='000000'))
           
            if i>=total_roll+6 :
                mysheet2[f'A{i}'].alignment= Alignment(horizontal='left', vertical='center') 
        
        for i in range(total_roll+15,total_roll+22):
            mysheet2[f'B{i}'].alignment= Alignment(horizontal='center', vertical='center')     
            mysheet2[f'B{i}'].border=Border(top=Side(style='thin',color='000000'),right=Side(style='thin',color='000000'),left=Side(style='thin',color='000000'),bottom=Side(style='thin',color='000000'))
            mysheet2[f'C{i}'].alignment= Alignment(horizontal='center', vertical='center')     
            mysheet2[f'C{i}'].border=Border(top=Side(style='thin',color='000000'),right=Side(style='thin',color='000000'),left=Side(style='thin',color='000000'),bottom=Side(style='thin',color='000000'))
            
        mysheet2[f'B{total_roll+15}'] = "COs"
        mysheet2[f'B{total_roll+15}'].font=Font(bold=True)
        mysheet2[f'C{total_roll+15}'] = "AL"
        mysheet2[f'C{total_roll+15}'].font=Font(bold=True)
        mysheet2[f'B{total_roll+16}'] = 'CO1'
        mysheet2[f'B{total_roll+17}'] = 'CO2'
        mysheet2[f'B{total_roll+18}'] = 'CO3'
        mysheet2[f'B{total_roll+19}'] = 'CO4'
        mysheet2[f'B{total_roll+20}'] = 'CO5'
        mysheet2[f'B{total_roll+21}'] = 'CO6' 
    
    def make_CA_Type_Quiz(mysheet,ca_array):
        temp=len(ca_array)
        # mysheet=sheet4
        
        mysheet.column_dimensions['B'].width =42
        mysheet['A1']="Roll No."
        mysheet['B1']="Name"
        
        if temp==1 :
            mysheet['C1']="Q1"
            mysheet['C2']="CO"+ca_array[0]+""
            myArr=['A','B', 'C']
        
        elif temp==2:
            mysheet['C1']="Q1"
            mysheet['D1']="Q2"
            
            mysheet['C2']="CO"+ca_array[0]+""
            mysheet['D2']="CO"+ca_array[1]+""
            myArr=['A','B', 'C', 'D']
            
        elif temp==3:
            mysheet['C1']="Q1"
            mysheet['D1']="Q2"
            mysheet['E1']="Q3"
            
            mysheet['C2']="CO"+ca_array[0]+""
            mysheet['D2']="CO"+ca_array[1]+""
            mysheet['E2']="CO"+ca_array[2]+""
            myArr=['A','B', 'C', 'D', 'E']
            
        elif temp==4:
            mysheet['C1']="Q1"
            mysheet['D1']="Q2"
            mysheet['E1']="Q3"
            mysheet['F1']="Q4"
            
            mysheet['C2']="CO"+ca_array[0]+""
            mysheet['D2']="CO"+ca_array[1]+""
            mysheet['E2']="CO"+ca_array[2]+""
            mysheet['F2']="CO"+ca_array[3]+""
            myArr=['A','B', 'C', 'D', 'E', 'F']
            
        elif temp==5:
            mysheet['C1']="Q1"
            mysheet['D1']="Q2"
            mysheet['E1']="Q3"
            mysheet['F1']="Q4"
            mysheet['G1']="Q5"
            
            mysheet['C2']="CO"+ca_array[0]+""
            mysheet['D2']="CO"+ca_array[1]+""
            mysheet['E2']="CO"+ca_array[2]+""
            mysheet['F2']="CO"+ca_array[3]+""
            mysheet['G2']="CO"+ca_array[4]+""
            myArr=['A','B', 'C', 'D', 'E', 'F', 'G']
            
        elif temp==6:
            mysheet['C1']="Q1"
            mysheet['D1']="Q2"
            mysheet['E1']="Q3"
            mysheet['F1']="Q4"
            mysheet['G1']="Q5"
            mysheet['H1']="Q6"
            
            mysheet['C2']="CO"+ca_array[0]+""
            mysheet['D2']="CO"+ca_array[1]+""
            mysheet['E2']="CO"+ca_array[2]+""
            mysheet['F2']="CO"+ca_array[3]+""
            mysheet['G2']="CO"+ca_array[4]+""
            mysheet['H2']="CO"+ca_array[5]+""
            myArr=['A','B', 'C', 'D', 'E', 'F', 'G','H']
            
        elif temp==7:
            mysheet['C1']="Q1"
            mysheet['D1']="Q2"
            mysheet['E1']="Q3"
            mysheet['F1']="Q4"
            mysheet['G1']="Q5"
            mysheet['H1']="Q6"
            mysheet['I1']="Q7"
            
            mysheet['C2']="CO"+ca_array[0]+""
            mysheet['D2']="CO"+ca_array[1]+""
            mysheet['E2']="CO"+ca_array[2]+""
            mysheet['F2']="CO"+ca_array[3]+""
            mysheet['G2']="CO"+ca_array[4]+""
            mysheet['H2']="CO"+ca_array[5]+""
            mysheet['I2']="CO"+ca_array[6]+""
            myArr=['A','B', 'C', 'D', 'E', 'F', 'G','H','I']
            
        elif temp==8:
            mysheet['C1']="Q1"
            mysheet['D1']="Q2"
            mysheet['E1']="Q3"
            mysheet['F1']="Q4"
            mysheet['G1']="Q5"
            mysheet['H1']="Q6"
            mysheet['I1']="Q7"
            mysheet['J1']="Q8"
            
            mysheet['C2']="CO"+ca_array[0]+""
            mysheet['D2']="CO"+ca_array[1]+""
            mysheet['E2']="CO"+ca_array[2]+""
            mysheet['F2']="CO"+ca_array[3]+""
            mysheet['G2']="CO"+ca_array[4]+""
            mysheet['H2']="CO"+ca_array[5]+""
            mysheet['I2']="CO"+ca_array[6]+""
            mysheet['J2']="CO"+ca_array[7]+""
            myArr=['A','B', 'C', 'D', 'E', 'F', 'G','H','I', 'J']
            
        elif temp==9:
            mysheet['C1']="Q1"
            mysheet['D1']="Q2"
            mysheet['E1']="Q3"
            mysheet['F1']="Q4"
            mysheet['G1']="Q5"
            mysheet['H1']="Q6"
            mysheet['I1']="Q7"
            mysheet['J1']="Q8"
            mysheet['K1']="Q9"
            
            mysheet['C2']="CO"+ca_array[0]+""
            mysheet['D2']="CO"+ca_array[1]+""
            mysheet['E2']="CO"+ca_array[2]+""
            mysheet['F2']="CO"+ca_array[3]+""
            mysheet['G2']="CO"+ca_array[4]+""
            mysheet['H2']="CO"+ca_array[5]+""
            mysheet['I2']="CO"+ca_array[6]+""
            mysheet['J2']="CO"+ca_array[7]+""
            mysheet['K2']="CO"+ca_array[8]+""
            myArr=['A','B', 'C', 'D', 'E', 'F', 'G','H','I', 'J','K']
            
        else:
            mysheet['C1']="Q1"
            mysheet['D1']="Q2"
            mysheet['E1']="Q3"
            mysheet['F1']="Q4"
            mysheet['G1']="Q5"
            mysheet['H1']="Q6"
            mysheet['I1']="Q7"
            mysheet['J1']="Q8"
            mysheet['K1']="Q9"
            mysheet['L1']="Q10"
            
            mysheet['C2']="CO"+ca_array[0]+""
            mysheet['D2']="CO"+ca_array[1]+""
            mysheet['E2']="CO"+ca_array[2]+""
            mysheet['F2']="CO"+ca_array[3]+""
            mysheet['G2']="CO"+ca_array[4]+""
            mysheet['H2']="CO"+ca_array[5]+""
            mysheet['I2']="CO"+ca_array[6]+""
            mysheet['J2']="CO"+ca_array[7]+""
            mysheet['K2']="CO"+ca_array[8]+""
            mysheet['L2']="CO"+ca_array[9]+"" 
            myArr=['A','B', 'C', 'D', 'E', 'F', 'G','H','I', 'J','K','L']
        
        
        for i in range(1,3):
            for col in  ['A','B', 'C', 'D', 'E', 'F', 'G','H','I', 'J','K','L'] :
                mysheet[f'{col}{i}'].font=Font(bold=True)
                
        mysheet.merge_cells('A2:B2')
        
        
        for i in range(1 ,total_roll+1):
            mysheet[f'A{i+2}']=i
        
        for i in range(total_roll+3,total_roll+11) :
            mysheet.merge_cells(f'A{i}:B{i}')
            
        for i in range(1,total_roll+10):
            for j,col in enumerate(myArr[:temp+2]) :
                mysheet[f'{col}{i}'].alignment= Alignment(horizontal='center', vertical='center')     
                mysheet[f'{col}{i}'].border=Border(top=Side(style='thin',color='000000'),right=Side(style='thin',color='000000'),left=Side(style='thin',color='000000'),bottom=Side(style='thin',color='000000'))
                if i>total_roll+2 :
                    mysheet[f'{col}{i}'].alignment= Alignment(horizontal='left', vertical='center')     
                
        for i in range(total_roll+3,total_roll+10):
            start_index=1
            for j, col in enumerate(myArr[start_index:temp+2],start=start_index+1) :
                mysheet[f'{col}{i}'].alignment= Alignment(horizontal='center', vertical='center')     
                mysheet[f'{col}{i}'].border=Border(top=Side(style='thin',color='000000'),right=Side(style='thin',color='000000'),left=Side(style='thin',color='000000'),bottom=Side(style='thin',color='000000'))
            if i==total_roll+9:
                mysheet[f'{col}{i}'].font=Font(bold=True)        
                
    
        mysheet[f'A{total_roll+3}']="Count(Attempted)"       
        mysheet[f'A{total_roll+4}']="Average Marks"
        
    
        mysheet[f'A{total_roll+5}']=f"Count(>={basic_values_temp[9]}%)"
        
        
        mysheet[f'A{total_roll+6}']=f"% Count(>={basic_values_temp[9]}% w.r.t appeared)"

        mysheet[f'A{total_roll+7}']="Count(>=Average Marks of class)"
        mysheet[f'A{total_roll+8}']="% Count(>=Average Marks of class w.r.t appeared)"
        
        mysheet[f'A{total_roll+9}']=f"AL(Based on >={basic_values_temp[9]}% Count) (All COs)"
        mysheet[f'A{total_roll+9}'].font=Font(bold=True)
        
        for i in range(total_roll+12,total_roll+19):
            mysheet[f'C{i}'].alignment= Alignment(horizontal='center', vertical='center')     
            mysheet[f'C{i}'].border=Border(top=Side(style='thin',color='000000'),right=Side(style='thin',color='000000'),left=Side(style='thin',color='000000'),bottom=Side(style='thin',color='000000'))
            mysheet[f'D{i}'].alignment= Alignment(horizontal='center', vertical='center')     
            mysheet[f'D{i}'].border=Border(top=Side(style='thin',color='000000'),right=Side(style='thin',color='000000'),left=Side(style='thin',color='000000'),bottom=Side(style='thin',color='000000'))
            
        mysheet[f'C{total_roll+12}'] = "COs"
        mysheet[f'C{total_roll+12}'].font=Font(bold=True)
        mysheet[f'D{total_roll+12}'] = "AL"
        mysheet[f'D{total_roll+12}'].font=Font(bold=True)
        mysheet[f'C{total_roll+13}'] = 'CO1'
        mysheet[f'C{total_roll+14}'] = 'CO2'
        mysheet[f'C{total_roll+15}'] = 'CO3'
        mysheet[f'C{total_roll+16}'] = 'CO4'
        mysheet[f'C{total_roll+17}'] = 'CO5'
        mysheet[f'C{total_roll+18}'] = 'CO6' 
    
    print(basic_values_temp[10])       
    if basic_values_temp[10]=="2":
        if basic_values_temp[11]=="Quiz":
            make_CA_Type_Quiz(sheet3,CA1_Co_arr_temp)
        else :
            make_CA_Type_PPT_NPTEL(sheet3,CA1_Co_arr_temp)
        print("Hello",basic_values_temp[10]) 
        if basic_values_temp[12]=="Quiz":
            make_CA_Type_Quiz(sheet4,CA2_Co_arr_temp)
        else :
            make_CA_Type_PPT_NPTEL(sheet4,CA2_Co_arr_temp)
    elif basic_values_temp[10]=="3":
        if basic_values_temp[11]=="Quiz":
            make_CA_Type_Quiz(sheet3,CA1_Co_arr_temp)
        else :
            make_CA_Type_PPT_NPTEL(sheet3,CA1_Co_arr_temp)
        print("Hi",basic_values_temp[10]) 
        if basic_values_temp[12]=="Quiz":
            make_CA_Type_Quiz(sheet4,CA2_Co_arr_temp)
        else :
            make_CA_Type_PPT_NPTEL(sheet4,CA2_Co_arr_temp) 
             
        if basic_values_temp[13]=="Quiz":
            make_CA_Type_Quiz(sheet7,CA3_Co_arr_temp)
        else :
            make_CA_Type_PPT_NPTEL(sheet7,CA3_Co_arr_temp) 
        print("Hi",basic_values_temp[13])   
    #<-----------Survey Template------------->
    sheet5.column_dimensions['B'].width =40
    sheet5.column_dimensions['C'].width =40
    sheet5.column_dimensions['F'].width =40
    
    sheet5['A1']="Sr. No."
    sheet5['B1']="Email Address"
    sheet5['C1']="Full name of Student"
    sheet5['D1']="Roll No."
    sheet5['E1']="Class"
    sheet5['F1']="Branch"
    sheet5['G1']="Q1"
    sheet5['H1']="Q2"
    sheet5['I1']="Q3"
    sheet5['J1']="Q4"
    sheet5['K1']="Q5"
    sheet5['L1']="Q6"
    
    for col in  ['A','B', 'C', 'D', 'E', 'F', 'G','H','I', 'J','K','L'] :
            sheet5[f'{col}1'].font=Font(bold=True)
            
    for i in range(1 ,total_roll+1):
        sheet5[f'A{i+1}']=i
        sheet5[f'E{i+1}']=""+basic_values_temp[7]+""
        sheet5[f'F{i+1}']=""+basic_values_temp[1]+""
        
    for i in range(1,total_roll+2):
        for col in ['A','B', 'C', 'D', 'E', 'F', 'G','H','I', 'J','K','L'] :
            sheet5[f'{col}{i}'].alignment= Alignment(horizontal='center', vertical='center')     
            sheet5[f'{col}{i}'].border=Border(top=Side(style='thin',color='000000'),right=Side(style='thin',color='000000'),left=Side(style='thin',color='000000'),bottom=Side(style='thin',color='000000'))
    
    sheet5[f'F{total_roll+4}']= 'Total' 
    sheet5[f'F{total_roll+4}'].font=Font(bold=True) 
    sheet5[f'F{total_roll+5}']= 'SA + A Count'
    sheet5[f'F{total_roll+5}'].font=Font(bold=True)
    sheet5[f'F{total_roll+6}']= 'SA + A Percentage' 
    sheet5[f'F{total_roll+6}'].font=Font(bold=True)
    sheet5[f'F{total_roll+7}']= 'CO Mapped' 
    sheet5[f'F{total_roll+7}'].font=Font(bold=True)
    sheet5[f'F{total_roll+8}']= 'AL'
    
    for i in range(total_roll+4,total_roll+9):
        for col in ['F','G','H','I','J','K','L'] :
            sheet5[f'{col}{i}'].alignment= Alignment(horizontal='center', vertical='center')     
            sheet5[f'{col}{i}'].border=Border(top=Side(style='thin',color='000000'),right=Side(style='thin',color='000000'),left=Side(style='thin',color='000000'),bottom=Side(style='thin',color='000000'))
            if col=="F":
                sheet5[f'F{i}'].alignment= Alignment(horizontal='left', vertical='center')     
    
            if i==total_roll+8:
                sheet5[f'{col}{i}'].font=Font(bold=True)
        
    sheet5[f'G{total_roll+7}']= 'CO1' 
    sheet5[f'H{total_roll+7}']= 'CO2' 
    sheet5[f'I{total_roll+7}']= 'CO3' 
    sheet5[f'J{total_roll+7}']= 'CO4' 
    sheet5[f'K{total_roll+7}']= 'CO5' 
    sheet5[f'L{total_roll+7}']= 'CO6'
    
    
    
    #<-----------------------Attainment--------------------->
    
    sheet6.column_dimensions['A'].width =16
    sheet6.column_dimensions['B'].width =25
    sheet6.column_dimensions['C'].width =25
    sheet6.column_dimensions['D'].width =25
    sheet6.column_dimensions['E'].width =25
    sheet6.column_dimensions['F'].width =34
    sheet6.column_dimensions['G'].width =25
    
    for i in range (1,9):
        sheet6.merge_cells(f"A{i}:H{i}")
        sheet6[f'A{i}'].font=Font(bold=True)
        
    
    for i in range (9,15):
        sheet6.merge_cells(f"B{i}:H{i}") 
         
    

    sheet6["A1"].value="Vivekanand Education Society's Institute of Technology"
    sheet6["A1"].alignment= Alignment(horizontal='center', vertical='center')     
    
    sheet6["A2"].value="Department of "+basic_values_temp[1]+""
    sheet6["A2"].alignment= Alignment(horizontal='center', vertical='center')     
    
    sheet6["A3"].value="Academic Year :"+basic_values_temp[5]+""
    sheet6["A3"].alignment= Alignment(horizontal='center', vertical='center')     
    
    sheet6["A5"].value="  Subject : "+basic_values_temp[4]+"                                                                                                                                                                       Class : "+basic_values_temp[7]+""
    sheet6["A5"].alignment= Alignment(horizontal='left', vertical='center')     
    
    sheet6["A6"].value="  Subject Teacher :"+basic_values_temp[6]+"                                                                                                                                                                Semester : "+basic_values_temp[3]+""
    sheet6["A6"].alignment= Alignment(horizontal='left', vertical='center')     
    
    
    sheet6['A8']='Course Outcomes(COs): Upon successful completion of this course, students will be able to:'
    sheet6['A8'].font=Font(bold=True)
    sheet6["A8"].alignment= Alignment(horizontal='left', vertical='center')     
    
    sheet6['A9'] ='CO1'
    sheet6['A10']='CO2'
    sheet6['A11']='CO3'
    sheet6['A12']='CO4'
    sheet6['A13']='CO5'
    sheet6['A14']='CO6'
    
    for i in range (9,15):
        sheet6[f'A{i}'].alignment= Alignment(horizontal='center', vertical='center')
        sheet6[f'B{i}'].alignment= Alignment(horizontal='left', vertical='center')         
    
    for i in range(9,15):
        for col in ['A','B', 'C', 'D', 'E', 'F', 'G','H']:
            sheet6[f'{col}{i}'].border=Border(top=Side(style='thin',color='000000'),right=Side(style='thin',color='000000'),left=Side(style='thin',color='000000'),bottom=Side(style='thin',color='000000'))  
                    
    sheet6.merge_cells("A15:H15")
    sheet6.merge_cells("A16:H16")
    
    sheet6['A16']='CO Rubrics Mapping'
    sheet6['A16'].alignment= Alignment(horizontal='center', vertical='center')     
    
    sheet6.merge_cells("A17:H17")
    
    sheet6.merge_cells("A18:A19")  
    sheet6.merge_cells("F18:F19")
    
    sheet6.merge_cells("B18:E18")
    sheet6.merge_cells("B19:D19") 
    
    for i in range(16,21):
        for col in ['A','B', 'C', 'D', 'E', 'F', 'G','H']:
            sheet6[f'{col}{i}'].font=Font(bold=True)
    
    for i in range(18,27):
        for col in ['A','B', 'C', 'D', 'E', 'F']:
            sheet6[f'{col}{i}'].border=Border(top=Side(style='thin',color='000000'),right=Side(style='thin',color='000000'),left=Side(style='thin',color='000000'),bottom=Side(style='thin',color='000000')) 
            sheet6[f'{col}{i}'].alignment= Alignment(horizontal='center', vertical='center')     
                
    sheet6['A18']='Assessment'
    sheet6['B18']='Direct Assessment' 
    sheet6['F18']='Indirect Assessment' 
    
    sheet6['B19']='Internal Assessment' 
    sheet6['E19']='External Assessment' 
    
    sheet6['A20']="CO's"
    sheet6['B20']="Mid Term Test"
    sheet6['C20']="Countinous Assessment"
    sheet6['D20']="Quiz"
    sheet6['E20']="ESE(TH)"
    sheet6['F20']="Course Exit Survey"
    
    sheet6['A21']='CO1'
    sheet6['A22']='CO2'
    sheet6['A23']='CO3'
    sheet6['A24']='CO4'
    sheet6['A25']='CO5'
    sheet6['A26']='CO6'
    
    sheet6.merge_cells("A27:H27")
    sheet6.merge_cells("A28:H28")
    sheet6.merge_cells("A29:H29")
    
    sheet6['A28']='CO Attainment (Level)'
    sheet6['A28'].alignment= Alignment(horizontal='center', vertical='center')     
    
    sheet6.merge_cells("A30:A31")
    sheet6.merge_cells("G30:G31")
    
    sheet6.merge_cells("B30:F30")
    sheet6.merge_cells("B31:D31")
    
    for i in range(28,33):
        for col in ['A','B', 'C', 'D', 'E', 'F', 'G','H']:
            sheet6[f'{col}{i}'].font=Font(bold=True)
    
    for i in range(30,39):
        for col in ['A','B', 'C', 'D', 'E', 'F', 'G']:
            sheet6[f'{col}{i}'].border=Border(top=Side(style='thin',color='000000'),right=Side(style='thin',color='000000'),left=Side(style='thin',color='000000'),bottom=Side(style='thin',color='000000')) 
            sheet6[f'{col}{i}'].alignment= Alignment(horizontal='center', vertical='center')     
       
    sheet6['A30']='Assessment'
    sheet6['B30']='Direct Assessment'
    sheet6['G30']='Indirect Assessment'
    
    sheet6['B31']='Internal Assessment'
    sheet6['E31']='External Assessment'
    sheet6['F31']='Attainment Level'
    
    sheet6['A32']="CO's"
    sheet6['B32']="Mid Term Test"
    sheet6['C32']="Countinous Assessment"
    sheet6['D32']="Quiz"
    sheet6['E32']="ESE(TH)"
    sheet6['F32']="70% (External) + 30% (Internal)"
    sheet6['G32']="Course Exit Survey"
    
    sheet6['A33']='CO1'
    sheet6['A34']='CO2'
    sheet6['A35']='CO3'
    sheet6['A36']='CO4'
    sheet6['A37']='CO5'
    sheet6['A38']='CO6'
    
    sheet6.merge_cells("A39:H39")
    sheet6.merge_cells("A40:H40")
    sheet6.merge_cells("A41:H41")
    
    for i in range(42,49):
        for col in ['C', 'D', 'E']:
            sheet6[f'{col}{i}'].border=Border(top=Side(style='thin',color='000000'),right=Side(style='thin',color='000000'),left=Side(style='thin',color='000000'),bottom=Side(style='thin',color='000000')) 
            sheet6[f'{col}{i}'].alignment= Alignment(horizontal='center', vertical='center')     
     
    sheet6['A40']='Final CO Attainment'
    sheet6['A40'].font=Font(bold=True)
    sheet6['A40'].alignment= Alignment(horizontal='center', vertical='center')     
    
    sheet6['C42']='Course Outcomes'
    sheet6['C42'].font=Font(bold=True)
    
    sheet6['D42']='Direct AL'
    sheet6['D42'].font=Font(bold=True)
    
    sheet6['E42']='Indirect AL'
    sheet6['E42'].font=Font(bold=True)
    
    sheet6['C43']='CO1'
    sheet6['C43'].font=Font(bold=True)
    
    sheet6['C44']='CO2'
    sheet6['C44'].font=Font(bold=True)
    
    sheet6['C45']='CO3'
    sheet6['C45'].font=Font(bold=True)
    
    sheet6['C46']='CO4'
    sheet6['C46'].font=Font(bold=True)
    
    sheet6['C47']='CO5'
    sheet6['C47'].font=Font(bold=True)
    
    sheet6['C48']='CO6'
    sheet6['C48'].font=Font(bold=True)
    
    
    workbook.save('C:/Users/bhush/Downloads/NewTemplate.xlsx')
